from openpyxl import load_workbook

from ggotaiorder.rpa.backup import BackupWriter
from ggotaiorder.rpa.models import RpaOrder


def _order(**kw):
    base = dict(
        order_detail_id=7, shop_key=3, shop_name="꽃집", channel="전화",
        customer_name="홍길동", customer_phone_number="010-0000-0000",
        product_name="장미다발", quantity=2, price=30000,
        delivery_at="2026-06-04 10:00", delivery_place="서울시 강남구",
        receiver_name="김철수", receiver_phone_number="010-1111-1111",
        ribbon_sender="보내는분", ribbon_congratulations="축 결혼",
        card_message="행복하세요",
    )
    base.update(kw)
    return RpaOrder(**base)


def test_backup_writes_xlsx_and_txt(tmp_path):
    writer = BackupWriter(tmp_path / "backups")
    xlsx_path, txt_path = writer.write(_order())

    assert xlsx_path.exists()
    assert txt_path.exists()

    ws = load_workbook(xlsx_path).active
    data_row = [c.value for c in ws[2]]      # 1행=헤더, 2행=값
    assert "장미다발" in data_row
    assert ws.cell(row=2, column=8).value == 2   # 8열=수량

    txt = txt_path.read_text(encoding="utf-8")
    assert "장미다발" in txt
    assert "김철수" in txt


def test_backup_creates_missing_dir(tmp_path):
    target = tmp_path / "nested" / "backups"
    writer = BackupWriter(target)
    writer.write(_order())
    assert target.exists()


def test_backup_includes_delivery_at_text(tmp_path):
    writer = BackupWriter(tmp_path / "backups")
    xlsx_path, txt_path = writer.write(_order(delivery_at_text="내일 오후 3시"))

    assert "내일 오후 3시" in txt_path.read_text(encoding="utf-8")
    # xlsx 에도 원문이 실린다(어느 열이든)
    ws = load_workbook(xlsx_path).active
    assert "내일 오후 3시" in [c.value for c in ws[2]]
    # 기존 열 위치 보존(수량=8열)
    assert ws.cell(row=2, column=8).value == 2


def test_receipt_renders_none_fields_as_dash(tmp_path):
    writer = BackupWriter(tmp_path / "backups")
    _, txt_path = writer.write(
        _order(delivery_place=None, ribbon_sender=None, card_message=None)
    )
    txt = txt_path.read_text(encoding="utf-8")
    assert "None" not in txt          # None이 그대로 출력되면 안 됨
    assert "배송지: -" in txt
